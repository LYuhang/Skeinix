"""Internal tabular I/O helpers for Platform MCP workflow execution.

Supported file types are detected from the extension: csv, tsv, jsonl, json,
xlsx. Addressing is uniform — rows by 1-based number/range, columns by name.
"""
from __future__ import annotations

import csv
import io
import json

from vibecanvas_api.agents.tools.decorator import ToolError
from vibecanvas_api.agents.tools.render import Rendered

_PREVIEW_ROWS = 5

_NO_WORKSPACE = ("no workspace is available — file operations require an active "
                 "workspace sandbox")


async def _session_call(coro):
    """Run a session fileop, translating an internal RuntimeError (no fileop pool)
    into a CLEAN no_workspace ToolError — never leak the raw backend string (matches
    the fs tools' per-call guard)."""
    try:
        return await coro
    except RuntimeError:
        raise ToolError("no_workspace", _NO_WORKSPACE)


def _ext(path: str) -> str:
    return path.rsplit(".", 1)[-1].lower() if "." in path else ""


# ── read ──────────────────────────────────────────────────────────────────

def _resolve_sheet(path: str, names: list[str], sheet: str) -> str:
    """Pick the target sheet, or raise a CLEAN ToolError. ``sheet=''`` + a single
    sheet → that sheet; ``''`` + multiple sheets → ``sheet_required`` (lists names);
    a named sheet must exist → else ``sheet_not_found`` (lists names)."""
    sheet = (sheet or "").strip()
    if sheet:
        if sheet in names:
            return sheet
        raise ToolError("sheet_not_found", f"sheet {sheet!r} not found in {path!r}; "
                        f"available sheets: {', '.join(names)}")
    if len(names) == 1:
        return names[0]
    raise ToolError("sheet_required", f"path {path!r} has multiple sheets "
                    f"({', '.join(names)}); specify one with the sheet argument")


def _xlsx_to_rows(data: bytes, path: str, sheet: str) -> tuple[list[dict], list[str]]:
    """Return (rows, columns). ``columns`` is the authoritative header (non-empty
    cells of the first row) — preserved even when there are zero data rows."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        raise ToolError("read_failed", f"could not read {path!r} as a spreadsheet")
    chosen = _resolve_sheet(path, wb.sheetnames, sheet)   # ToolError propagates (not swallowed)
    try:
        it = wb[chosen].iter_rows(values_only=True)
        try:
            headers = ["" if h is None else str(h) for h in next(it)]
        except StopIteration:
            return [], []
        cols = [h for h in headers if h]
        out = []
        for vals in it:
            out.append({headers[i]: (vals[i] if i < len(vals) else None)
                        for i in range(len(headers)) if headers[i]})
        return out, cols
    except Exception:
        raise ToolError("read_failed", f"could not read {path!r} as a spreadsheet")


def _text_to_rows(text: str, ext: str) -> tuple[list[dict], list[str]] | None:
    """Return (rows, columns) or None for an unrecognized format. For csv/tsv the
    columns are the header (kept even with zero data rows); for jsonl/json there is no
    separate header, so columns are the union of the rows' keys."""
    if ext in ("csv", "tsv"):
        delim = "\t" if ext == "tsv" else ","
        reader = csv.DictReader(io.StringIO(text), delimiter=delim)
        rows = list(reader)
        return rows, [c for c in (reader.fieldnames or []) if c is not None]
    if ext == "jsonl":
        rows = [json.loads(ln) for ln in text.splitlines() if ln.strip()]
        return rows, _all_columns(rows)
    if ext == "json":
        obj = json.loads(text or "[]")
        if isinstance(obj, dict) and "rows" in obj:
            rows = obj["rows"]
        elif isinstance(obj, list):
            rows = obj
        else:
            return None
        return rows, _all_columns(rows)
    return None


def _raise_read_error(res: dict, path: str) -> None:
    """Map an in-sandbox read failure to a CLEAN ToolError (never the raw cause)."""
    err = res.get("error") or ""
    if err == "not_found":
        raise ToolError("path_not_found", f"path {path!r} does not exist")
    if err == "path_outside_roots":
        raise ToolError("invalid_path", f"path {path!r} is outside the allowed roots")
    raise ToolError("read_failed", f"could not read {path!r}")


async def _read_rows(path: str, session, sheet: str = "") -> tuple[list[dict], list[str]]:
    """Read a file into (rows, columns) IN the sandbox. ``columns`` is the table's
    authoritative header (preserved even for a zero-row table). ``sheet`` selects a
    worksheet for xlsx (ignored for text formats). Raises ToolError (clean message)
    on a missing/unreadable/non-tabular file or an ambiguous/unknown sheet."""
    ext = _ext(path)
    if ext in ("xlsx", "xlsm"):
        res = await _session_call(session.read_bytes(path))
        if not res.get("ok"):
            _raise_read_error(res, path)
        return _xlsx_to_rows(res["data"], path, sheet)
    res = await _session_call(session.read_file(path))
    if not res.get("ok"):
        _raise_read_error(res, path)
    if res.get("kind") != "text":
        raise ToolError("not_text", f"path {path!r} is not a text table — read it with read_file")
    try:
        parsed = _text_to_rows(res.get("content", ""), ext)
    except (ValueError, TypeError):
        raise ToolError("parse_failed", f"could not parse {path!r}")
    if parsed is None:
        raise ToolError("not_tabular",
                        f"path {path!r} is not a recognized table (csv/tsv/jsonl/json/xlsx); "
                        f"read it with read_file")
    return parsed


# ── write ─────────────────────────────────────────────────────────────────

def _rows_to_text(rows: list[dict], ext: str, columns: list[str]) -> str:
    if ext in ("csv", "tsv"):
        delim = "\t" if ext == "tsv" else ","
        buf = io.StringIO()
        w = csv.DictWriter(buf, fieldnames=columns, delimiter=delim, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
        return buf.getvalue()
    if ext == "jsonl":
        return "\n".join(json.dumps(r, ensure_ascii=False) for r in rows)
    if ext == "json":
        return json.dumps(rows, ensure_ascii=False, indent=2)
    raise ValueError(f"unsupported write format: .{ext}")


def _rows_to_xlsx(rows: list[dict], columns: list[str], sheet: str,
                  base: bytes | None, path: str) -> bytes:
    """Serialize rows into an xlsx workbook. With ``base`` (existing file bytes),
    load it and replace ONLY the target sheet — preserving the others (read-modify-
    write); without it, create a new single-sheet workbook. Raises ToolError when the
    target sheet is ambiguous (multiple existing sheets, none named). A named sheet
    that does not yet exist is created."""
    import openpyxl
    name = (sheet or "").strip()
    if base is not None:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(base))
        except Exception:
            raise ToolError("read_failed", f"could not read {path!r} as a spreadsheet")
        if not name:
            if len(wb.sheetnames) == 1:
                name = wb.sheetnames[0]
            else:
                raise ToolError("sheet_required", f"path {path!r} has multiple sheets "
                                f"({', '.join(wb.sheetnames)}); specify one with the sheet argument")
        if name in wb.sheetnames:                       # replace in place (keep position)
            idx = wb.sheetnames.index(name)
            del wb[name]
            ws = wb.create_sheet(name, idx)
        else:                                           # a new sheet in the existing book
            ws = wb.create_sheet(name)
    else:                                               # brand-new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = name or "Sheet1"
    ws.append(columns)
    for r in rows:
        ws.append([r.get(c) for c in columns])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _write_rows(path: str, rows: list[dict], session, sheet: str = "") -> int:
    """Write rows to a file (by extension) IN the sandbox — text formats via
    ``session.write_file``, xlsx via ``session.write_bytes``. For xlsx the existing
    workbook is loaded first so only the target ``sheet`` is replaced (others kept).
    Returns bytes written; raises ToolError (clean message) on failure."""
    columns = _all_columns(rows)            # union of keys, so a column added to any row is kept
    ext = _ext(path)
    if ext in ("xlsx", "xlsm"):
        existing = await _session_call(session.read_bytes(path))
        base = existing["data"] if existing.get("ok") else None
        payload = _rows_to_xlsx(rows, columns, sheet, base, path)
        res = await _session_call(session.write_bytes(path, payload))
    else:
        try:
            text = _rows_to_text(rows, ext, columns)
        except ValueError:
            raise ToolError("bad_format", f"path {path!r} has an unsupported table format "
                            f"(.{ext}) — use csv / tsv / jsonl / json / xlsx")
        res = await _session_call(session.write_file(path, text))
    if not res.get("ok"):
        err = res.get("error") or ""
        if err == "path_outside_roots":
            raise ToolError("invalid_path", f"path {path!r} is outside the allowed roots")
        raise ToolError("write_failed", f"could not write {path!r}")
    return res.get("bytes") or 0


def _data_render(raw: dict, ctx) -> Rendered:
    """Shared render for the read-style data tools: content = the data itself (rows
    as jsonl, else pretty JSON) so a large table is offloaded by the decorator; the
    summary is the abstract; content_type drives compaction."""
    ct = raw.get("content_type") or "application/json"
    data = raw.get("data")
    if isinstance(data, list):
        content = "\n".join(json.dumps(r, ensure_ascii=False, default=str) for r in data)
    else:
        content = json.dumps(data, ensure_ascii=False, default=str, indent=2)
    return Rendered(content=content, content_type=ct, abstract=raw.get("summary", ""))


# ── addressing: rows by 1-based number/range, columns by name ─────────────

def _row_index_set(spec: str, n: int) -> list[int]:
    """Parse a 1-based rows spec to 0-based indices. "" = all; "5" = one row;
    "2:10" = an inclusive range. Raises a CLEAN ``invalid_rows`` ToolError on a
    malformed spec (never leaks the raw int() ValueError)."""
    spec = (spec or "").strip()
    if not spec:
        return list(range(n))
    try:
        if ":" in spec:
            lo, _, hi = spec.partition(":")
            start = int(lo) if lo.strip() else 1
            end = int(hi) if hi.strip() else n
            return [i - 1 for i in range(start, end + 1) if 1 <= i <= n]
        i = int(spec)
        return [i - 1] if 1 <= i <= n else []
    except ValueError:
        raise ToolError("invalid_rows", f"invalid rows spec {spec!r} — use a 1-based row "
                        'number ("5") or an inclusive range ("2:10")')


def _col_list(spec: str, all_cols: list[str]) -> list[str]:
    spec = (spec or "").strip()
    if not spec:
        return all_cols
    return [c.strip() for c in spec.split(",") if c.strip()]


def _select_columns(spec: str, all_cols: list[str], path: str) -> list[str]:
    """Column selection for the READ tools: "" = all columns; otherwise every named
    column must exist, else a CLEAN ``column_not_found`` ToolError listing what is
    available."""
    cols = _col_list(spec, all_cols)
    missing = [c for c in cols if c not in all_cols]
    if missing:
        raise ToolError("column_not_found",
                        f"column(s) {', '.join(repr(m) for m in missing)} not found in "
                        f"{path!r}; available columns: {', '.join(all_cols) or '(none)'}")
    return cols


def _all_columns(rows: list[dict]) -> list[str]:
    cols: list[str] = []
    for r in rows:
        for k in r.keys():
            if k not in cols:
                cols.append(k)
    return cols
