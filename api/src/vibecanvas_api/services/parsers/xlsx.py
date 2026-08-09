"""XLSX parser with sheet and row-range citation metadata."""
from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from .archive import validate_office_archive
from .base import EmptyDocumentError, ParsedSegment, ParseError, Parser


ROWS_PER_SEGMENT = 100
MAX_SHEETS = 200
MAX_ROWS = 200_000


class XlsxParser(Parser):
    def parse(self, blob: bytes) -> list[ParsedSegment]:
        validate_office_archive(blob, required_prefix="xl/")
        try:
            workbook = load_workbook(BytesIO(blob), read_only=True, data_only=True)
        except Exception as exc:
            raise ParseError(f"Cannot parse xlsx: {exc}") from exc
        if len(workbook.worksheets) > MAX_SHEETS:
            raise ParseError("xlsx contains too many sheets")
        segments: list[ParsedSegment] = []
        total_rows = 0
        try:
            for sheet in workbook.worksheets:
                batch: list[str] = []
                batch_start = 1
                for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    total_rows += 1
                    if total_rows > MAX_ROWS:
                        raise ParseError("xlsx contains too many rows")
                    values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                    if values:
                        if not batch:
                            batch_start = row_number
                        batch.append(" | ".join(values))
                    if len(batch) >= ROWS_PER_SEGMENT:
                        segments.append(ParsedSegment(
                            text="\n".join(batch),
                            metadata={"sheet": sheet.title, "row_start": batch_start, "row_end": row_number},
                        ))
                        batch = []
                if batch:
                    segments.append(ParsedSegment(
                        text="\n".join(batch),
                        metadata={"sheet": sheet.title, "row_start": batch_start, "row_end": row_number},
                    ))
        finally:
            workbook.close()
        if not segments:
            raise EmptyDocumentError("xlsx has no readable cells")
        return segments
