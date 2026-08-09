# -*- coding: utf-8 -*-
"""Local structured-file I/O for TableReadNode / TableWriteNode — CSV, JSONL, Excel.

Self-contained in the engine: stdlib ``csv``/``json`` for CSV/JSONL and
``openpyxl`` for Excel, with no dependency on a separate application tree.
Paths handed here are already real (a ``/run`` path mapped to the per-run dir, or
any local path) — these helpers just open them.
"""
from __future__ import annotations

import csv
import json
import os
from typing import Any


_EXT_FORMAT = {
    ".csv": "csv",
    ".json": "jsonl",
    ".jsonl": "jsonl",
    ".ndjson": "jsonl",
    ".xlsx": "excel",
    ".xls": "excel",
}


def detect_format(path: str) -> str:
    """Infer ``csv`` / ``jsonl`` / ``excel`` from the file extension."""
    ext = os.path.splitext(path)[1].lower()
    fmt = _EXT_FORMAT.get(ext)
    if fmt is None:
        raise ValueError(
            f"Cannot detect format from '{path}'. Supported: "
            ".csv, .json/.jsonl/.ndjson, .xlsx/.xls. Set file_format explicitly.")
    return fmt


# --------------------------------------------------------------------------- #
# Read                                                                         #
# --------------------------------------------------------------------------- #
def read_rows(path: str, fmt: str, sheet_name: str | None = None) -> list[dict]:
    """Read all rows from a local file as a list of dicts."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")
    if fmt == "csv":
        return _read_csv(path)
    if fmt == "jsonl":
        return _read_jsonl(path)
    if fmt == "excel":
        return _read_excel(path, sheet_name or None)
    raise ValueError(f"Unsupported file_format '{fmt}' (csv | jsonl | excel).")


def read_headers(path: str, fmt: str, sheet_name: str | None = None) -> list[str]:
    """Read ONLY the column headers of an existing file (cheap — first row/record).

    Used to verify append-mode schema correspondence without loading the whole
    file. Returns ``[]`` for an empty file; raises ``FileNotFoundError`` when the
    path is missing and ``ValueError`` for an unsupported format.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")
    if fmt == "csv":
        with open(path, "r", newline="", encoding="utf-8-sig") as fp:
            for row in csv.reader(fp):
                return [str(h) for h in row]
            return []
    if fmt == "jsonl":
        with open(path, "r", encoding="utf-8") as fp:
            for line in fp:
                line = line.strip()
                if line:
                    obj = json.loads(line)
                    return list(obj.keys()) if isinstance(obj, dict) else []
            return []
    if fmt == "excel":
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        it = ws.iter_rows(values_only=True)
        try:
            headers = [str(h) if h is not None else "" for h in next(it)]
        except StopIteration:
            headers = []
        wb.close()
        return headers
    raise ValueError(f"Unsupported file_format '{fmt}' (csv | jsonl | excel).")


def _read_csv(path: str) -> list[dict]:
    with open(path, "r", newline="", encoding="utf-8-sig") as fp:
        return [dict(row) for row in csv.DictReader(fp)]


def _read_jsonl(path: str) -> list[dict]:
    rows = []
    with open(path, "r", encoding="utf-8") as fp:
        for line in fp:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def _read_excel(path: str, sheet_name: str | None) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    it = ws.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else "" for h in next(it)]
    except StopIteration:
        wb.close()
        return []
    rows = []
    for values in it:
        if values is None or all(v is None for v in values):
            continue
        rows.append({headers[i]: (values[i] if i < len(values) else None)
                     for i in range(len(headers))})
    wb.close()
    return rows


# --------------------------------------------------------------------------- #
# Write                                                                        #
# --------------------------------------------------------------------------- #
def write_rows(path: str, fmt: str, rows: list[dict], headers: list[str],
               sheet_name: str | None = None, append: bool = False) -> None:
    """Write rows to a local file. ``append`` adds to an existing file (CSV/JSONL)
    or sheet (Excel); otherwise the file is (re)created. Parent dirs are created."""
    parent = os.path.dirname(path)
    if parent:
        os.makedirs(parent, exist_ok=True)
    if fmt == "csv":
        _write_csv(path, rows, headers, append)
    elif fmt == "jsonl":
        _write_jsonl(path, rows, append)
    elif fmt == "excel":
        _write_excel(path, rows, headers, sheet_name or None, append)
    else:
        raise ValueError(f"Unsupported file_format '{fmt}' (csv | jsonl | excel).")


def _write_csv(path: str, rows: list[dict], headers: list[str], append: bool) -> None:
    exists = append and os.path.exists(path)
    with open(path, "a" if append else "w", newline="", encoding="utf-8") as fp:
        writer = csv.DictWriter(fp, fieldnames=headers, extrasaction="ignore")
        if not exists:
            writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _write_jsonl(path: str, rows: list[dict], append: bool) -> None:
    with open(path, "a" if append else "w", encoding="utf-8") as fp:
        for row in rows:
            fp.write(json.dumps(row, ensure_ascii=False) + "\n")


def _write_excel(path: str, rows: list[dict], headers: list[str],
                 sheet_name: str | None, append: bool) -> None:
    from openpyxl import Workbook, load_workbook
    if append and os.path.exists(path):
        wb = load_workbook(path)
        ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active
    else:
        wb = Workbook()
        ws = wb.active
        if sheet_name:
            ws.title = sheet_name
        ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    wb.save(path)
    wb.close()


# --------------------------------------------------------------------------- #
# Schema inference (informational — the TableReadNode `schema` output)          #
# --------------------------------------------------------------------------- #
def infer_schema(rows: list[dict], greedy_parse: bool = True, max_depth: int = 3) -> dict:
    """Infer a coarse JSON Schema for a row from the first row's values."""
    if not rows:
        return {"type": "object", "properties": {}}
    props = {k: _infer_value(v, 0, max_depth, greedy_parse) for k, v in rows[0].items()}
    return {"type": "object", "properties": props}


def _infer_value(val: Any, depth: int, max_depth: int, greedy_parse: bool) -> dict:
    if greedy_parse and isinstance(val, str):
        s = val.strip()
        if s and s[0] in "[{":
            try:
                val = json.loads(s)
            except (ValueError, TypeError):
                pass
    if isinstance(val, bool):
        return {"type": "boolean"}
    if isinstance(val, int):
        return {"type": "integer"}
    if isinstance(val, float):
        return {"type": "number"}
    if val is None:
        return {"type": "null"}
    if isinstance(val, list):
        if depth >= max_depth or not val:
            return {"type": "array"}
        return {"type": "array", "items": _infer_value(val[0], depth + 1, max_depth, greedy_parse)}
    if isinstance(val, dict):
        if depth >= max_depth:
            return {"type": "object"}
        return {"type": "object",
                "properties": {k: _infer_value(v, depth + 1, max_depth, greedy_parse)
                               for k, v in val.items()}}
    return {"type": "string"}
