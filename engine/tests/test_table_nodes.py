# -*- coding: utf-8 -*-
"""TableReadNode / TableWriteNode — local-only (CSV/JSONL/Excel), identity paths.

Verifies: write-then-read round-trip for each format against a real directory;
under the identity-bind model a node opens the path it is given DIRECTLY (inside
the sandbox ``/run`` / ``/mount`` are available at those literal paths, so a
logical path IS the real file — no resolution step). Also offset/limit and that
legacy cross-application coupling is absent.
"""
import os

import pytest

from vibecanvas_engine.nodes.table_read import TableReadNode
from vibecanvas_engine.nodes.table_write import TableWriteNode
from vibecanvas_engine.nodes import table_io

_ROWS = [{"name": "a", "n": 1}, {"name": "b", "n": 2}, {"name": "c", "n": 3}]


def _call(node, inputs, extra):
    # @safe_call_with_args wraps the node return in an envelope; the engine
    # consumes ``["output"]``. Mirror that here.
    res = node(inputs, {}, extra=extra)
    assert res["status"] == "success", res.get("error_message")
    return res["output"]


def _writer(path, fmt=None, mode="overwrite"):
    n = TableWriteNode.__new__(TableWriteNode)
    cfg = {"file_path": path, "write_mode": mode}
    if fmt:
        cfg["file_format"] = fmt
    n.node_config = cfg
    return n


def _reader(path, fmt=None, **extra_cfg):
    n = TableReadNode.__new__(TableReadNode)
    cfg = {"file_path": path}
    if fmt:
        cfg["file_format"] = fmt
    cfg.update(extra_cfg)
    n.node_config = cfg
    return n


@pytest.mark.parametrize("fname", ["data.csv", "data.jsonl", "data.xlsx"])
def test_write_then_read_roundtrip(tmp_path, fname):
    run = {"run_id": "r1", "run_dir": str(tmp_path)}
    # The real path the bind exposes inside the sandbox (identity).
    path = os.path.join(str(tmp_path), fname)

    out = _call(_writer(path), {"rows": _ROWS}, run)
    # The path is opened/echoed verbatim, file actually created there.
    assert out["file_path"] == path
    assert os.path.isfile(out["file_path"])
    assert out["rows_written"] == 3

    res = _call(_reader(path), {}, run)
    assert res["row_count"] == 3
    assert res["headers"] == ["name", "n"]
    assert [r["name"] for r in res["rows"]] == ["a", "b", "c"]
    assert res["schema"]["type"] == "object" and "name" in res["schema"]["properties"]


def test_offset_and_limit(tmp_path):
    run = {"run_id": "r1", "run_dir": str(tmp_path)}
    path = os.path.join(str(tmp_path), "d.jsonl")
    _call(_writer(path), {"rows": _ROWS}, run)
    res = _call(_reader(path, offset=1, limit=1), {}, run)
    assert res["row_count"] == 1 and res["rows"][0]["name"] == "b"


def test_append_mode(tmp_path):
    run = {"run_id": "r1", "run_dir": str(tmp_path)}
    path = os.path.join(str(tmp_path), "log.csv")
    _call(_writer(path, mode="overwrite"), {"rows": _ROWS[:1]}, run)
    _call(_writer(path, mode="append"), {"rows": _ROWS[1:]}, run)
    res = _call(_reader(path), {}, run)
    assert res["row_count"] == 3


def test_absolute_local_path_passthrough(tmp_path):
    # Any absolute path is used as-is (no extra needed).
    p = str(tmp_path / "abs.jsonl")
    _call(_writer(p), {"rows": _ROWS}, {})
    assert os.path.isfile(p)
    res = _call(_reader(p), {}, {})
    assert res["row_count"] == 3


def test_table_nodes_have_no_legacy_application_coupling():
    assert TableReadNode.CONFIG_SCHEMA["properties"]["file_format"]["enum"] == [
        "auto", "csv", "jsonl", "excel"
    ]
    assert TableWriteNode.CONFIG_SCHEMA["properties"]["file_format"]["enum"] == [
        "auto", "csv", "jsonl", "excel"
    ]
    import inspect
    for mod in (TableReadNode, TableWriteNode, table_io):
        src = inspect.getsource(inspect.getmodule(mod))
        assert "sys.path.insert" not in src

    # Both nodes are thread-bridged so they receive `extra` (run_dir).
    assert TableReadNode.REQUIRES_THREAD_BRIDGE is True
    assert TableWriteNode.REQUIRES_THREAD_BRIDGE is True


def test_sheet_name_interpolates_placeholders(tmp_path):
    # sheet_name supports {{var}} the same way file_path does — for both
    # write (the sheet it creates) and read (the sheet it targets).
    run = {"run_id": "r1", "run_dir": str(tmp_path)}
    inputs = {"month": "jan", "rows": _ROWS}
    path = os.path.join(str(tmp_path), "book.xlsx")

    writer = _writer(path, fmt="excel")
    writer.node_config["sheet_name"] = "data_{{month}}"
    _call(writer, inputs, run)

    # The interpolated sheet name was actually used on disk.
    import openpyxl
    wb = openpyxl.load_workbook(path)
    assert "data_jan" in wb.sheetnames

    # Reading via the same {{var}} sheet_name resolves to that sheet.
    reader = _reader(path, fmt="excel", sheet_name="data_{{month}}")
    res = _call(reader, {"month": "jan"}, run)
    assert res["row_count"] == 3
    assert [r["name"] for r in res["rows"]] == ["a", "b", "c"]


def test_sheet_name_plain_and_empty_passthrough(tmp_path):
    # No placeholders: a plain name passes through unchanged; empty stays empty.
    run = {"run_id": "r1", "run_dir": str(tmp_path)}
    plain = os.path.join(str(tmp_path), "plain.xlsx")
    writer = _writer(plain, fmt="excel")
    writer.node_config["sheet_name"] = "Sheet1"
    _call(writer, {"rows": _ROWS}, run)
    import openpyxl
    wb = openpyxl.load_workbook(plain)
    assert "Sheet1" in wb.sheetnames

    # Empty sheet_name still works (round-trips through the default sheet).
    empty = os.path.join(str(tmp_path), "empty.xlsx")
    _call(_writer(empty, fmt="excel"), {"rows": _ROWS}, run)
    res = _call(_reader(empty, fmt="excel"), {}, run)
    assert res["row_count"] == 3


def test_detect_format_rejects_unknown():
    with pytest.raises(ValueError):
        table_io.detect_format("data.parquet")
